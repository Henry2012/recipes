#! /usr/bin/python
#-*- coding:utf-8 -*-

"""
Author: Qiqun Han
File: xlsx.recipes.py
Description: this program gives methods to process Microsoft Excel.
Creation: 2014-1-13
Revision: 2014-1-13
"""

from openpyxl import load_workbook
from openpyxl import Workbook

#===============================================================================
# 读取已经有的excel文件
#===============================================================================

# wb = load_workbook(filename=r'../stat/email_analysis/NO ANSWERS - test set - trials - with corporate emails.xlsx')
# 
# # 返回该xlsx文件下所有的sheets
# print wb.get_sheet_names()
# 
# # 读取某一个sheet，注意sheet名是case-sensitive.
# working_sheet = wb.get_sheet_by_name(name='Sheet1')
# 
# # 读取具体某个单元格的值
# print working_sheet.cell('D18').value  # D18
# 
# # 在某个sheet里读取多个单元格的值
# for each in working_sheet.range('B1:B100'):
#     print each
    
#===============================================================================
# 如何写入数据
# 需要预先知道写入多少行多少列，不然会相当耗内存
#===============================================================================

wb = Workbook()

# 方法1
ws = wb.get_active_sheet()

# 方法2
# ws2 = wb.create_sheet()
# ws3 = web.create_sheet(0)

# 检验已经创建的sheets
print wb.get_sheet_names()

original = 0
for i in xrange(10):
    for j in xrange(10):
        ws.cell(row=i, column=j).value = original
        original += 1
wb.save("../io/test_writing_into_xlsx.xlsx")